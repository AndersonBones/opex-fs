from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle
import os
import datetime


file_path = "result.xlsx"
wb = load_workbook(file_path)

# grab the active worksheet
ws = wb['Opex']




def set_font():
    for row in ws["A1:W100"]:
        for cell in row:
            cell.font = Font(name='Montserrat',
                size=8,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')

def number_format():
    for row in ws["B4:Q90"]:
        for cell in row:
            cell.number_format  = numbers.FORMAT_NUMBER_COMMA_SEPARATED2


    


def adjust_column():
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.9
        ws.column_dimensions[column_letter].width = adjusted_width



def get_subtotal_rows():
    rows = []
    for index, cell in enumerate(ws["A4:A90"]):
        index_account = index+4
        if cell[0].value is not None:
            if cell[0].value[0].isalpha() == True:
                rows.append(index_account)

    return rows




def set_subtotais():
    rows = get_subtotal_rows()

    print(rows)

    for cell_index in rows:
        for row in ws[f"A{cell_index}:Q{cell_index}"]:
            for cell in row:
        
                cell.fill = PatternFill("solid", fgColor="1A7753")
                                    
                cell.font = Font(name='Montserrat',
                                size=9,
                                bold=True,
                                italic=False,
                                vertAlign=None,
                                underline='none',
                                strike=False,
                                color='ffffff')
                    
        
                


def insert_row():
    rows = get_subtotal_rows()
    for index, row in enumerate(rows):
        ws.insert_rows(rows[index]+len(rows)-len(rows)+index, amount=1)


insert_row()

adjust_column()
set_font()
number_format()
set_subtotais()


def set_datetime_header():
    safra_month = [4,5,6,7,8,9,10,11,12,1,2,3]

    ws["A2"].value = ""
    ws["A3"].value = ""

    ws.merge_cells("A1:A2")

    if ws["B1"].value == "Budget" and ws["C1"].value == "Forecast":
        ws.merge_cells("B1:B2")
        ws.merge_cells("C1:C2")


    
    ws["D3"] = datetime.datetime(year, 4, 1)
    ws["D3"].number_format = "mmm-yy"


    
    for index, cell in ws["D2:O3"]:
        year = 0
        if index == 0: 
            year = int(cell.value)
        
        if index == 1:
            cell = datetime.datetime(year, 4, 1)

set_datetime_header()
wb.save(file_path)




