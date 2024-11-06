from openpyxl import load_workbook
from openpyxl.styles import numbers
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle, alignment
from openpyxl.styles.borders import BORDER_THIN
from openpyxl.formatting.rule import IconSet, FormatObject
from openpyxl.formatting.rule import Rule

from openpyxl.chart import BarChart, Series, Reference

import os
import datetime





class Formatting:
    def __init__(self, output_path_opex) -> None:
        self.output_path_opex = output_path_opex


        self.wb = load_workbook(self.output_path_opex)
        

        # grab the active worksheet
        self.ws = self.wb['Opex']

        self.ws.sheet_view.showGridLines = False

        self.colors = {
            "yellow":"F7D21E",
            "green":"1B7754",
            "dark_green":"0E3A29",
            "gray-500":"D9D9D9",
            "gray-600":"D9D9D9",
            "white":"ffffff",
            "dark_gray":"3A3838"
        }
        



    def set_font(self): # set font style 
        for row in self.ws["A1:W100"]:
            for cell in row:
                cell.font = Font(name='Montserrat',
                    size=9,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='000000')
                
                cell.alignment = alignment.Alignment(vertical="center")


    def number_format_style(self): # set number format 
        FORMAT_NUMBER_COMMA_SEPARATED = '#,##0'
        font_color = ''

        for row in self.ws["B4:T90"]:
            for cell in row:
                cell.number_format  = FORMAT_NUMBER_COMMA_SEPARATED

                if (type(cell.value) == int or type(cell.value) == float) and cell.value < 0:
                    font_color = 'FF0101'
                   
                else:
                    font_color = '3A3838'
                 


                cell.font = Font(name='Montserrat', # set font style row
                                size=9,
                                bold=False,
                                italic=False,
                                vertAlign=None,
                                underline='none',
                                strike=False,
                                color=font_color)
                    


                


    def get_subtotal_rows(self): # get subtotal rows
        rows = []
        for index, cell in enumerate(self.ws["A4:A90"]):
            index_account = index+4
            if cell[0].value is not None:
                if cell[0].value[0].isalpha() == True:
                    rows.append(index_account)

        return rows




    def set_subtotais_style(self): # set subtotal style rows
        rows = self.get_subtotal_rows()
        bg_color = self.colors['green']
        font_color = self.colors['white']

        for cell_index in rows: # looping in rows
            for row in self.ws[f"A{cell_index}:T{cell_index}"]:
                for index, cell in enumerate(row):
                    
                    if cell_index > 4:
                        bg_color = self.colors['gray-600']
                        font_color = self.colors['dark_gray']
                    else:
                        bg_color = self.colors['green']
                        font_color = self.colors['white']
                        
                    cell.fill = PatternFill("solid", fgColor=bg_color) # set background row
                                        
                    cell.font = Font(name='Montserrat', # set font style row
                                size=9,
                                bold=True,
                                italic=False,
                                vertAlign=None,
                                underline='none',
                                strike=False,
                                color=font_color)
    


    def set_header_style(self):
        for cell in self.ws['A1:T1'][0]:
            cell.fill = PatternFill("solid", fgColor="0E3A29") # set background row

            cell.font = Font(name='Montserrat', # set font style row
                                    size=9,
                                    bold=True,
                                    italic=False,
                                    vertAlign=None,
                                    underline='none',
                                    strike=False,
                                    color='ffffff')
            cell.alignment = alignment.Alignment(horizontal="center", vertical="center")
            cell.border = None

        for cell in self.ws["D2:O2"][0]:
            cell.fill = PatternFill("solid", fgColor="0E3A29") # set background row

            cell.font = Font(name='Montserrat', # set font style row
                                    size=9,
                                    bold=True,
                                    italic=False,
                                    vertAlign=None,
                                    underline='none',
                                    strike=False,
                                    color='ffffff')
            
            cell.border = None
            cell.alignment = alignment.Alignment(horizontal="center", vertical="center")


    def insert_empty_row(self): # insert empty row after last row in subtotal table
        rows = self.get_subtotal_rows()
        for index, row in enumerate(rows):
           self.ws.insert_rows(rows[index]+len(rows)-len(rows)+index, amount=1)



    def set_safra_header(self): # set safra header: months and years
        months = [4, 5, 6, 7, 8,9,10, 11, 12, 1, 2, 3]
        years = []

        # get years 
        for cell in self.ws["D2:O2"][0]:
            years.append(cell.value)
        
        # set month
        for index, cell in enumerate(self.ws["D3:O3"][0]):
            cell.value = datetime.datetime(int(years[index]), months[index], 1) # set datetime in cell value
            cell.number_format = "mmm-yy" # set date format
        
    


    def merge_cells_header(self): # merge cells by status Forecast or Realizado
        
        self.ws.delete_rows(idx=2, amount=1)

        # delete values in cells
        self.ws["A2"].value = ""
        self.ws["A3"].value = ""

        self.ws.merge_cells("A1:A2") # merge conta header title

        # merge Forecast and Budget cells
        if self.ws["B1"].value == "Budget" and self.ws["C1"].value == "Forecast":
            self.ws.merge_cells("B1:B2")


            self.ws.merge_cells("C1:C2")
        

        self.ws.merge_cells("P1:P2")
        self.ws.merge_cells("Q1:Q2")
        self.ws.merge_cells("R1:R2")
        self.ws.merge_cells("S1:S2")
        self.ws.merge_cells("T1:T2")


    def set_forecast_realizado_header(self): # set forecast or realizado by month 
        for index, cell in enumerate(self.ws["D1:O1"][0]):
            year = self.ws["D2:O2"][0][index].value.year # get year value from row years
            month = self.ws["D2:O2"][0][index].value.month # get month value from row months

            # set label forecast or realizado by date
            if (year > datetime.date.today().year and month <= datetime.date.today().month) or (year <= datetime.date.today().year and month >= datetime.date.today().month):
                self.ws["D1:O2"][0][index].value = "Forecast"
            else:
                self.ws["D1:O2"][0][index].value = "Realizado"
            


        
        # merge styling
        forecast_cells_coordinate = []
        realizado_cells_coordinate = []

        # get Realizado and Forecast in row
        for index, cell in enumerate(self.ws["D1:O1"][0]): 
            if cell.value == "Forecast":
                forecast_cells_coordinate.append(cell)
      
            if cell.value == "Realizado":
                realizado_cells_coordinate.append(cell)

        # merge Forecast cells and Realizado cells
        self.ws.merge_cells(f"{forecast_cells_coordinate[0].coordinate}:{forecast_cells_coordinate[-1].coordinate}")
        self.ws.merge_cells(f"{realizado_cells_coordinate[0].coordinate}:{realizado_cells_coordinate[-1].coordinate}")

        
    def set_icons(self):

        first = FormatObject(type='num', val=-99999)
        second = FormatObject(type='num', val=99999)

        ytd_iconset = IconSet(iconSet='3Symbols', cfvo=[first, second], showValue=None, percent=None, reverse=True)
        ytd_rule = Rule(type='iconSet', iconSet=ytd_iconset)

        self.ws.conditional_formatting.add("R4:R60", ytd_rule)


        bdg_fct_iconset = IconSet(iconSet='3Arrows', cfvo=[first, second], showValue=None, percent=None, reverse=None)
        bdg_fct_rule = Rule(type='iconSet', iconSet=bdg_fct_iconset)

        self.ws.conditional_formatting.add("S4:S60", bdg_fct_rule)

        
        


    def save(self):
        self.wb.save(filename=self.output_path_opex)

        # os.popen("Opex.xlsx")

    def run(self):
        self.set_font()

        # set formats
        self.insert_empty_row()
        self.set_safra_header()
        self.merge_cells_header()
        self.set_forecast_realizado_header()


        #style
        self.set_header_style()
        self.number_format_style()
        self.set_subtotais_style()

        self.set_icons()
        
        # save file
        self.save() 